"""Local file knowledge base indexer.

Reads documents from a directory (PDFs, text files, images) and builds
a searchable in-memory index. Students should extend this with better
chunking, embeddings, or a vector store as their project evolves.

Supported formats:
    - .pdf  (requires PyPDF2 or pdfplumber)
    - .txt, .md
    - .xlsx (reads first sheet as text)
    - .jpg, .jpeg, .png (stores metadata only — use OCR tool for content)
"""

from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


class KnowledgeBaseIndexer:
    """Index local files into a searchable knowledge base.

    This is a starter implementation using simple text chunking.
    Replace with embeddings + vector store for production quality.
    """

    def __init__(self, knowledge_dir: str, chunk_size: int = 500, chunk_overlap: int = 50):
        self.knowledge_dir = Path(knowledge_dir)
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.documents: list[dict[str, Any]] = []

    def index_all(self) -> int:
        """Index all supported files in the knowledge directory."""
        if not self.knowledge_dir.exists():
            logger.warning(f"Knowledge directory not found: {self.knowledge_dir}")
            return 0

        self.documents = []
        count = 0
        for file_path in self.knowledge_dir.iterdir():
            if file_path.is_file():
                try:
                    self.index_file(str(file_path))
                    count += 1
                except Exception as e:
                    logger.error(f"Failed to index {file_path.name}: {e}")

        logger.info(f"Indexed {count} files, {len(self.documents)} chunks total")
        return count

    def index_file(self, file_path: str) -> None:
        """Index a single file into the knowledge base."""
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".pdf":
            text = self._read_pdf(path)
        elif suffix in (".txt", ".md"):
            text = path.read_text(encoding="utf-8", errors="ignore")
        elif suffix == ".xlsx":
            text = self._read_xlsx(path)
        elif suffix in (".jpg", ".jpeg", ".png"):
            self.documents.append({
                "title": path.name,
                "content": f"[Image file: {path.name}. Use the OCR tool to extract text from this image.]",
                "source": str(path),
                "type": "image",
            })
            return
        else:
            logger.debug(f"Skipping unsupported file type: {path.name}")
            return

        chunks = self._chunk_text(text, source=path.name)
        self.documents.extend(chunks)

    def _read_pdf(self, path: Path) -> str:
        """Extract text from a PDF file."""
        try:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
            return "\n\n".join(text_parts)
        except ImportError:
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(str(path))
                text_parts = []
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
                return "\n\n".join(text_parts)
            except ImportError:
                logger.warning(
                    f"No PDF library available. Install pdfplumber or PyPDF2. "
                    f"Skipping: {path.name}"
                )
                return ""

    def _read_xlsx(self, path: Path) -> str:
        """Extract text from an Excel file."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            text_parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        text_parts.append(row_text)
            return "\n".join(text_parts)
        except ImportError:
            logger.warning(f"openpyxl not installed. Skipping: {path.name}")
            return ""

    def _chunk_text(self, text: str, source: str) -> list[dict[str, Any]]:
        """Split text into overlapping chunks for search."""
        if not text.strip():
            return []

        chunks = []
        words = text.split()
        i = 0
        chunk_idx = 0

        while i < len(words):
            chunk_words = words[i : i + self.chunk_size]
            chunk_text = " ".join(chunk_words)

            chunks.append({
                "title": f"{source} (chunk {chunk_idx + 1})",
                "content": chunk_text,
                "source": source,
                "type": "text",
                "chunk_index": chunk_idx,
            })

            i += self.chunk_size - self.chunk_overlap
            chunk_idx += 1

        return chunks
